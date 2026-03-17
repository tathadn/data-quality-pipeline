"""
Data Quality Auditor — implements the full audit procedure from skills/data-auditor/SKILL.md
Input:  examples/raw-data/sample_sales.csv
Output: examples/audit-reports/
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend — REQUIRED
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import ListedColormap
from reportlab.lib.pagesizes import letter, landscape as ls_pagesize
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate, NextPageTemplate,
    Paragraph, Spacer, PageBreak, KeepTogether,
    Table, TableStyle, Image as RLImage,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils
from datetime import datetime
import os
import re
import tempfile
import shutil

# ── CONFIG ────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE = os.path.join(BASE_DIR, "examples", "raw-data", "building_permits.csv")
OUTPUT_DIR = os.path.join(BASE_DIR, "examples", "audit-reports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── COLOR PALETTE (from SKILL.md) ────────────────────────────────────────────
COLORS = {
    'primary':     '#1a56db',
    'accent':      '#10b981',
    'warning':     '#f59e0b',
    'danger':      '#ef4444',
    'purple':      '#8b5cf6',
    'gray_dark':   '#374151',
    'gray_medium': '#6b7285',
    'gray_light':  '#f3f4f6',
    'white':       '#ffffff',
}
SEVERITY_COLORS = {'High': '#ef4444', 'Medium': '#f59e0b', 'Low': '#10b981'}
CHART_PALETTE   = ['#1a56db', '#10b981', '#f59e0b', '#ef4444',
                   '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16']

plt.rcParams.update({
    'figure.facecolor': 'white',
    'axes.facecolor':   'white',
    'axes.edgecolor':   '#e5e7eb',
    'axes.grid':        True,
    'grid.alpha':       0.3,
    'grid.color':       '#e5e7eb',
    'font.family':      'sans-serif',
    'font.size':        10,
    'axes.titlesize':   13,
    'axes.titleweight': 'bold',
    'axes.labelsize':   10,
    'figure.dpi':       150,
})

# ── HELPERS ───────────────────────────────────────────────────────────────────
def sev_score(s):
    return {'High': 3, 'Medium': 2, 'Low': 1}.get(s, 0)

def classify_casing(v):
    s = str(v).strip()
    if not s or s == 'nan':
        return None
    if s.isupper():     return 'ALL_CAPS'
    if s.islower():     return 'all_lower'
    if s.istitle():     return 'Title_Case'
    return 'Mixed'

_DATE_PATTERNS = [
    (r'^\d{1,2}/\d{1,2}/\d{4}$',  'MM/DD/YYYY'),
    (r'^\d{4}-\d{2}-\d{2}$',      'YYYY-MM-DD'),
    (r'^\d{2}-\d{2}-\d{4}$',      'DD-MM-YYYY'),
    (r'^\d{1,2}-\d{1,2}-\d{4}$',  'MM-DD-YYYY'),
    (r'^\d{2}/\d{2}/\d{4}$',      'DD/MM/YYYY'),
]

def detect_date_formats(series):
    found = {}
    for val in series.dropna().astype(str).str.strip():
        for pattern, name in _DATE_PATTERNS:
            if re.match(pattern, val):
                found[name] = found.get(name, 0) + 1
                break
    return found

def is_date_column(series):
    sample = series.dropna().astype(str).str.strip().head(60)
    if len(sample) == 0:
        return False
    hits = sum(
        1 for v in sample
        if any(re.match(p, v) for p, _ in _DATE_PATTERNS)
    )
    return hits / len(sample) > 0.5

def parse_mixed_date(val):
    s = str(val).strip()
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def try_numeric(series):
    """Strip $/, commas and try to cast to float. Returns (parsed, pct_success)."""
    def _p(v):
        try:
            return float(str(v).strip().replace('$', '').replace(',', ''))
        except Exception:
            return np.nan
    parsed = series.apply(_p)
    n_non_null = series.notna().sum()
    pct = parsed.notna().sum() / n_non_null if n_non_null else 0
    return parsed, pct

# ── STEP 1: LOAD & INSPECT ────────────────────────────────────────────────────
print(f"\nLoading {INPUT_FILE} ...")
def _read_csv(*args, **kwargs):
    for enc in ('utf-8', 'latin-1', 'cp1252'):
        try:
            return pd.read_csv(*args, encoding=enc, **kwargs)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {INPUT_FILE} with utf-8, latin-1, or cp1252")

df_raw = _read_csv(INPUT_FILE, dtype=str, keep_default_na=False)
df     = _read_csv(INPUT_FILE)                   # pandas-inferred types

total_rows = len(df_raw)
total_cols = len(df_raw.columns)
dataset_name = os.path.basename(INPUT_FILE).rsplit('.', 1)[0]
today = datetime.now().strftime('%Y-%m-%d')

print(f"  {total_rows} rows × {total_cols} columns")
print(f"  Columns: {list(df_raw.columns)}")

# Identify column categories
numeric_cols_pandas   = df.select_dtypes(include=[np.number]).columns.tolist()
date_columns          = [c for c in df_raw.columns if is_date_column(df_raw[c])]
# Try to recover numeric cols that pandas stored as object (e.g. revenue)
extra_numeric = {}
for col in df_raw.columns:
    if col in numeric_cols_pandas or col in date_columns:
        continue
    parsed, pct = try_numeric(df_raw[col])
    if pct > 0.60:
        extra_numeric[col] = parsed

all_numeric_for_analysis = {c: df[c] for c in numeric_cols_pandas}
all_numeric_for_analysis.update(extra_numeric)

print(f"  Numeric (pandas): {numeric_cols_pandas}")
print(f"  Numeric (recovered): {list(extra_numeric.keys())}")
print(f"  Date columns: {date_columns}")

issues = []   # {column, check_type, severity, finding, affected_rows, percentage, example_values}

# ── CHECK 1: MISSING VALUES ───────────────────────────────────────────────────
missing_stats = {}
for col in df_raw.columns:
    s = df_raw[col]
    count = (s.str.strip() == '').sum()
    pct   = count / total_rows * 100
    missing_stats[col] = {'count': int(count), 'pct': pct}
    if count > 0:
        sev = 'High' if pct > 20 else ('Medium' if pct >= 5 else 'Low')
        examples = df_raw[col][s.str.strip() != ''].head(2).tolist()
        issues.append(dict(
            column=col, check_type='Missing Values', severity=sev,
            finding=f'{count} missing values ({pct:.1f}%)',
            affected_rows=int(count), percentage=round(pct, 2),
            example_values=str(examples),
        ))

# ── CHECK 2: DUPLICATE ROWS ───────────────────────────────────────────────────
dup_count = int(df_raw.duplicated().sum())
dup_pct   = dup_count / total_rows * 100
dup_sev   = 'High' if dup_pct > 5 else ('Medium' if dup_pct >= 1 else 'Low')
if dup_count > 0:
    issues.append(dict(
        column='ALL', check_type='Duplicate Rows', severity=dup_sev,
        finding=f'{dup_count} exact duplicate rows ({dup_pct:.1f}%)',
        affected_rows=dup_count, percentage=round(dup_pct, 2),
        example_values='N/A',
    ))

# ── CHECK 3: DATA TYPE VALIDATION ────────────────────────────────────────────
type_mismatch_pcts = []
for col in df_raw.columns:
    s = df_raw[col]
    non_empty = s[s.str.strip() != '']
    if len(non_empty) < 5:
        continue
    if col in date_columns:
        continue

    # Check: column parseable as numeric but has $ / comma formatting
    parsed, numeric_pct = try_numeric(non_empty)
    if numeric_pct > 0.60:
        str_formatted = non_empty[non_empty.str.strip().str.match(r'^\$[\d,]+\.?\d*$')]
        str_pct = len(str_formatted) / len(non_empty) * 100
        if str_pct > 5:
            sev = 'High' if str_pct > 20 else 'Medium'
            type_mismatch_pcts.append(str_pct)
            issues.append(dict(
                column=col, check_type='Type Validation', severity=sev,
                finding=f'{str_pct:.1f}% of values stored as currency strings (e.g. "$1,200")',
                affected_rows=int(str_pct * len(non_empty) / 100),
                percentage=round(str_pct, 2),
                example_values=str(str_formatted.head(3).tolist()),
            ))
        continue

    # Check: quantity-like integer column that has floats
    is_int_col = col in ('quantity',) or 'qty' in col.lower() or 'count' in col.lower()
    if is_int_col:
        float_vals = non_empty[non_empty.str.strip().str.match(r'^\d+\.\d+$')]
        if len(float_vals) > 0:
            fp = len(float_vals) / len(non_empty) * 100
            type_mismatch_pcts.append(fp)
            issues.append(dict(
                column=col, check_type='Type Validation', severity='Low',
                finding=f'{len(float_vals)} float values in expected-integer column ({fp:.1f}%)',
                affected_rows=len(float_vals), percentage=round(fp, 2),
                example_values=str(float_vals.head(3).tolist()),
            ))

avg_type_mismatch = np.mean(type_mismatch_pcts) if type_mismatch_pcts else 0.0

# ── CHECK 4: OUTLIER DETECTION ────────────────────────────────────────────────
outlier_stats = {}
for col, series in all_numeric_for_analysis.items():
    data = pd.to_numeric(series, errors='coerce').dropna()
    if len(data) < 10:
        continue
    q1, q3 = data.quantile(0.25), data.quantile(0.75)
    iqr = q3 - q1
    lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    outliers = data[(data < lo) | (data > hi)]
    pct = len(outliers) / len(data) * 100
    outlier_stats[col] = {'count': len(outliers), 'pct': pct, 'lower': lo, 'upper': hi}
    if len(outliers) > 0:
        sev = 'High' if pct > 10 else ('Medium' if pct >= 3 else 'Low')
        issues.append(dict(
            column=col, check_type='Outlier Detection', severity=sev,
            finding=f'{len(outliers)} outliers ({pct:.1f}%), valid range [{lo:.2f}, {hi:.2f}]',
            affected_rows=len(outliers), percentage=round(pct, 2),
            example_values=str(outliers.head(3).tolist()),
        ))

# ── CHECK 5: DATE FORMAT CONSISTENCY ─────────────────────────────────────────
for col in date_columns:
    fmt_map = detect_date_formats(df_raw[col])
    n_fmts  = len(fmt_map)
    sev = 'High' if n_fmts > 2 else ('Medium' if n_fmts == 2 else 'Low')
    issues.append(dict(
        column=col, check_type='Date Format Consistency', severity=sev,
        finding=f'{n_fmts} date format(s) detected: {", ".join(fmt_map.keys())}',
        affected_rows=total_rows, percentage=100.0,
        example_values=str(df_raw[col].head(3).tolist()),
    ))

# ── CHECK 6: TEXT CASING INCONSISTENCIES ─────────────────────────────────────
casing_pcts = []
for col in df_raw.columns:
    s = df_raw[col]
    non_empty = s[s.str.strip() != '']
    if len(non_empty) < 10 or col in date_columns:
        continue
    # Skip numeric-ish columns
    _, num_pct = try_numeric(non_empty)
    if num_pct > 0.5:
        continue
    casings = non_empty.apply(classify_casing).dropna()
    if len(casings) == 0:
        continue
    dist = casings.value_counts(normalize=True)
    dominant_pct = dist.iloc[0] * 100
    if dominant_pct < 80:
        bad_pct = 100 - dominant_pct
        casing_pcts.append(bad_pct)
        issues.append(dict(
            column=col, check_type='Casing Inconsistency', severity='Medium',
            finding=f'Dominant casing covers only {dominant_pct:.1f}%: {dict(casings.value_counts())}',
            affected_rows=int(bad_pct * len(casings) / 100), percentage=round(bad_pct, 2),
            example_values=str(non_empty.head(3).tolist()),
        ))

# ── CHECK 7: WHITESPACE ISSUES ────────────────────────────────────────────────
ws_pcts = []
for col in df_raw.columns:
    s = df_raw[col]
    non_empty = s[s.str.strip() != '']
    if len(non_empty) == 0:
        continue
    lt_mask     = non_empty != non_empty.str.strip()
    dbl_mask    = non_empty.str.contains(r'  +', regex=True) & ~lt_mask
    ws_count    = int(lt_mask.sum() + dbl_mask.sum())
    ws_pct      = ws_count / total_rows * 100
    if ws_count > 0:
        ws_pcts.append(ws_pct)
        sev = 'Medium' if ws_pct >= 10 else 'Low'
        issues.append(dict(
            column=col, check_type='Whitespace Issues', severity=sev,
            finding=f'{ws_count} values with leading/trailing or double spaces ({ws_pct:.1f}%)',
            affected_rows=ws_count, percentage=round(ws_pct, 2),
            example_values=str(non_empty[lt_mask].head(2).tolist()),
        ))

# ── CHECK 8: COLUMN NAME QUALITY ─────────────────────────────────────────────
unnamed_cols = [c for c in df_raw.columns if 'Unnamed' in str(c) or re.match(r'^col\d+$', str(c))]
bad_name_cols = [c for c in df_raw.columns
                 if c not in unnamed_cols and re.search(r'[^a-zA-Z0-9_]', str(c))]
if unnamed_cols:
    issues.append(dict(
        column=', '.join(unnamed_cols), check_type='Column Name Quality', severity='Medium',
        finding=f'{len(unnamed_cols)} unnamed/auto-generated column(s)',
        affected_rows=total_rows, percentage=100.0, example_values=str(unnamed_cols),
    ))
if bad_name_cols:
    issues.append(dict(
        column=', '.join(bad_name_cols), check_type='Column Name Quality', severity='Low',
        finding=f'{len(bad_name_cols)} column(s) with spaces/special chars in name',
        affected_rows=total_rows, percentage=100.0, example_values=str(bad_name_cols),
    ))

print(f"\n  {len(issues)} issues found across 8 checks")

# ── STEP 3: HEALTH SCORE ─────────────────────────────────────────────────────
avg_missing_pct   = np.mean([v['pct'] for v in missing_stats.values()])

# Extra penalty for columns >50% missing: 4 pts each, capped at 40
n_high_miss       = sum(1 for v in missing_stats.values() if v['pct'] > 50)
high_miss_penalty = min(40, n_high_miss * 4)
completeness      = max(0.0, 100 - avg_missing_pct - high_miss_penalty)

uniqueness        = max(0.0, 100 - dup_pct)
type_consistency  = max(0.0, 100 - avg_type_mismatch)
avg_outlier_pct   = np.mean([v['pct'] for v in outlier_stats.values()]) if outlier_stats else 0.0
outlier_reason    = max(0.0, 100 - avg_outlier_pct)

date_fmt_issues   = [i for i in issues if i['check_type'] == 'Date Format Consistency']
if date_fmt_issues:
    worst_fmt = max(sev_score(i['severity']) for i in date_fmt_issues)
    format_consist = max(0.0, 100 - worst_fmt * 25)
else:
    format_consist = 100.0

avg_text_pct      = np.mean(casing_pcts + ws_pcts) if (casing_pcts or ws_pcts) else 0.0
text_quality      = max(0.0, 100 - avg_text_pct)

# Weights: Completeness raised to 35%, Outlier/Format/Text trimmed to compensate
health_score = round(
    completeness     * 0.35 +
    uniqueness       * 0.15 +
    type_consistency * 0.20 +
    outlier_reason   * 0.12 +
    format_consist   * 0.10 +
    text_quality     * 0.08,
    1
)
health_rating = (
    'Excellent' if health_score >= 90 else
    'Good'      if health_score >= 70 else
    'Fair'      if health_score >= 50 else
    'Poor'
)

category_scores = {
    'Completeness':          completeness,
    'Uniqueness':            uniqueness,
    'Type Consistency':      type_consistency,
    'Outlier Reasonability': outlier_reason,
    'Format Consistency':    format_consist,
    'Text Quality':          text_quality,
}
cat_weights = {
    'Completeness': 35, 'Uniqueness': 15, 'Type Consistency': 20,
    'Outlier Reasonability': 12, 'Format Consistency': 10, 'Text Quality': 8,
}

high_issues   = sum(1 for i in issues if i['severity'] == 'High')
medium_issues = sum(1 for i in issues if i['severity'] == 'Medium')
low_issues    = sum(1 for i in issues if i['severity'] == 'Low')

print(f"\n  Health Score : {health_score}/100  ({health_rating})")
for cat, s in category_scores.items():
    print(f"    {cat:<24} {s:.1f}")
print(f"  Issues: {high_issues} High | {medium_issues} Medium | {low_issues} Low")

# ── STEP 4: GENERATE CHARTS ───────────────────────────────────────────────────
# Save chart images to a persistent subdir so ReportLab can always read them
charts_dir = os.path.join(OUTPUT_DIR, '_charts')
os.makedirs(charts_dir, exist_ok=True)
print(f"\n  Generating charts → {charts_dir} ...")

# ── Figure 1: Health Score Gauge ──────────────────────────────────────────────
gauge_path = os.path.join(charts_dir, 'gauge.png')
fig, ax = plt.subplots(figsize=(4, 4))
gauge_color = (COLORS['primary'] if health_score >= 70
               else COLORS['warning'] if health_score >= 50
               else COLORS['danger'])
ax.pie(
    [health_score, 100 - health_score],
    colors=[gauge_color, '#e5e7eb'],
    startangle=90, counterclock=False,
    wedgeprops={'width': 0.3, 'edgecolor': 'white', 'linewidth': 2},
)
ax.text(0,  0.05, f'{health_score:.0f}', fontsize=36, fontweight='bold',
        ha='center', va='center', color=COLORS['gray_dark'])
ax.text(0, -0.15, 'out of 100', fontsize=10,
        ha='center', va='center', color=COLORS['gray_medium'])
ax.text(0, -0.32, health_rating, fontsize=12, fontweight='bold',
        ha='center', va='center', color=gauge_color)
ax.set_title('Data Health Score', fontsize=14, fontweight='bold', pad=20)
plt.savefig(gauge_path, bbox_inches='tight', dpi=150)
plt.close()
print("    [1/6] gauge done")

# ── Figure 2: Missing Values Bar Chart ───────────────────────────────────────
missing_path = None
cols_with_missing = {k: v for k, v in missing_stats.items() if v['count'] > 0}
if cols_with_missing:
    missing_path = os.path.join(charts_dir, 'missing.png')
    sorted_miss = sorted(cols_with_missing.items(), key=lambda x: x[1]['pct'], reverse=True)
    col_names = [x[0] for x in sorted_miss]
    pcts      = [x[1]['pct'] for x in sorted_miss]
    bar_colors = [
        COLORS['danger']  if p > 20 else
        COLORS['warning'] if p >= 5 else
        COLORS['accent']
        for p in pcts
    ]
    fig, ax = plt.subplots(figsize=(8, max(4, len(col_names) * 0.55)))
    bars = ax.barh(col_names, pcts, color=bar_colors, edgecolor='white', height=0.6)
    ax.axvline(20, color=COLORS['danger'],  linestyle='--', linewidth=1,
               alpha=0.7, label='High threshold (20%)')
    ax.axvline(5,  color=COLORS['warning'], linestyle='--', linewidth=1,
               alpha=0.7, label='Medium threshold (5%)')
    for bar, pct in zip(bars, pcts):
        ax.text(bar.get_width() + 0.4,
                bar.get_y() + bar.get_height() / 2,
                f'{pct:.1f}%', va='center', ha='left', fontsize=9)
    ax.set_xlabel('Missing Values (%)')
    ax.set_title('Missing Values by Column', fontsize=14, fontweight='bold')
    ax.set_xlim(0, max(pcts) * 1.25)
    ax.legend(loc='lower right', fontsize=8)
    plt.tight_layout()
    plt.savefig(missing_path, bbox_inches='tight', dpi=150)
    plt.close()
print("    [2/6] missing-values bar chart done")

# ── Figure 3: Health Score Breakdown Pie ─────────────────────────────────────
pie_path = os.path.join(charts_dir, 'pie.png')
cat_labels = list(cat_weights.keys())
weighted_contributions = [category_scores[c] * cat_weights[c] / 100 for c in cat_labels]
fig, ax = plt.subplots(figsize=(7, 6))
wedges, texts, autotexts = ax.pie(
    weighted_contributions,
    labels=cat_labels,
    colors=CHART_PALETTE[:6],
    autopct='%1.0f%%',
    startangle=90,
    explode=[0.03] * 6,
    textprops={'fontsize': 9},
)
for at in autotexts:
    at.set_fontsize(8)
ax.set_title('Health Score Breakdown by Category', fontsize=14, fontweight='bold')
plt.savefig(pie_path, bbox_inches='tight', dpi=150)
plt.close()
print("    [3/6] score-breakdown pie done")

# ── Figure 4: Issue Severity Heatmap ─────────────────────────────────────────
heatmap_path = os.path.join(charts_dir, 'heatmap.png')
check_types  = [
    'Missing Values', 'Duplicate Rows', 'Type Validation',
    'Outlier Detection', 'Date Format Consistency',
    'Casing Inconsistency', 'Whitespace Issues', 'Column Name Quality',
]
columns_to_show = list(df_raw.columns)
if len(columns_to_show) > 20:
    col_issue_counts = {c: sum(1 for i in issues if c in i['column']) for c in columns_to_show}
    columns_to_show = sorted(col_issue_counts, key=col_issue_counts.get, reverse=True)[:20]

sev_matrix = np.zeros((len(check_types), len(columns_to_show)))
sev_labels  = np.full((len(check_types), len(columns_to_show)), '', dtype=object)

for issue in issues:
    ci = next((idx for idx, ct in enumerate(check_types) if ct == issue['check_type']), None)
    if ci is None:
        continue
    for cj, col in enumerate(columns_to_show):
        if col in issue['column'] or issue['column'] == 'ALL':
            v = sev_score(issue['severity'])
            if v > sev_matrix[ci, cj]:
                sev_matrix[ci, cj] = v
                sev_labels[ci, cj] = issue['severity'][0]  # H / M / L

sev_cmap = ListedColormap(['#f3f4f6', '#d1fae5', '#fef3c7', '#fee2e2'])
fig, ax = plt.subplots(figsize=(max(10, len(columns_to_show) * 0.75), 5))
sns.heatmap(
    sev_matrix, ax=ax, cmap=sev_cmap, vmin=0, vmax=3,
    xticklabels=columns_to_show, yticklabels=check_types,
    linewidths=1, linecolor='white', cbar=False,
    annot=sev_labels, fmt='s',
    annot_kws={'fontsize': 9, 'fontweight': 'bold'},
)
ax.set_title('Issue Severity Heatmap', fontsize=14, fontweight='bold')
plt.xticks(rotation=45, ha='right', fontsize=8)
plt.yticks(rotation=0, fontsize=8)
plt.tight_layout()
plt.savefig(heatmap_path, bbox_inches='tight', dpi=150)
plt.close()
print("    [4/6] severity heatmap done")

# ── Figure 5: Distribution Histograms ────────────────────────────────────────
hist_path = None
# Build a clean numeric frame: pandas numeric cols + recovered numeric cols
numeric_frame = pd.DataFrame()
for col in numeric_cols_pandas:
    numeric_frame[col] = df[col]
for col, parsed in extra_numeric.items():
    if col not in numeric_frame:
        numeric_frame[col] = parsed
hist_cols = list(numeric_frame.columns)[:6]

if hist_cols:
    hist_path = os.path.join(charts_dir, 'histograms.png')
    n = len(hist_cols)
    fig, axes = plt.subplots(nrows=2, ncols=3, figsize=(12, 7))
    axes = axes.flatten()
    for i, col in enumerate(hist_cols):
        ax  = axes[i]
        data = numeric_frame[col].dropna()
        ax.hist(data, bins=30, color=CHART_PALETTE[i % len(CHART_PALETTE)],
                edgecolor='white', alpha=0.85)
        q1, median, q3 = data.quantile([0.25, 0.5, 0.75])
        ax.axvline(median, color=COLORS['danger'], linestyle='--',
                   linewidth=1.5, label=f'Median: {median:.1f}')
        iqr = q3 - q1
        ax.axvline(q1 - 1.5 * iqr, color=COLORS['warning'],
                   linestyle=':', linewidth=1, alpha=0.7)
        ax.axvline(q3 + 1.5 * iqr, color=COLORS['warning'],
                   linestyle=':', linewidth=1, alpha=0.7)
        ax.set_title(col, fontsize=11, fontweight='bold')
        ax.legend(fontsize=7)
    for j in range(n, len(axes)):
        axes[j].set_visible(False)
    fig.suptitle('Value Distributions (Numeric Columns)',
                 fontsize=14, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(hist_path, bbox_inches='tight', dpi=150)
    plt.close()
print("    [5/6] distribution histograms done")

# ── Figure 6: Trend Lines ─────────────────────────────────────────────────────
trend_path = None
primary_date_col = date_columns[0] if date_columns else None
if primary_date_col and hist_cols:
    trend_cols = hist_cols[:3]
    df_trend   = pd.DataFrame({c: numeric_frame[c] for c in trend_cols})
    df_trend['_date'] = df_raw[primary_date_col].apply(parse_mixed_date)
    df_trend = df_trend.dropna(subset=['_date']).sort_values('_date')

    if len(df_trend) > 5:
        trend_path = os.path.join(charts_dir, 'trends.png')
        n_sub = len(trend_cols)
        fig, axes = plt.subplots(nrows=n_sub, ncols=1,
                                 figsize=(10, 3.5 * n_sub), sharex=True)
        if n_sub == 1:
            axes = [axes]
        for i, col in enumerate(trend_cols):
            ax = axes[i]
            x  = df_trend['_date'].values
            y  = df_trend[col].values
            ax.plot(x, y, color=CHART_PALETTE[i], alpha=0.4, linewidth=0.8)
            window  = max(7, len(df_trend) // 20)
            rolling = df_trend[col].rolling(window=window, center=True).mean()
            ax.plot(x, rolling.values, color=CHART_PALETTE[i], linewidth=2,
                    label=f'{window}-period avg')
            ax.set_ylabel(col, fontsize=10)
            ax.legend(fontsize=8)
        axes[-1].set_xlabel('Date')
        fig.suptitle('Trends Over Time', fontsize=14, fontweight='bold', y=1.01)
        plt.tight_layout()
        plt.savefig(trend_path, bbox_inches='tight', dpi=150)
        plt.close()
print("    [6/6] trend lines done")

# ── STEP 5: PDF REPORT ────────────────────────────────────────────────────────
pdf_filename = f"audit_report_{dataset_name}_{today}.pdf"
pdf_path     = os.path.join(OUTPUT_DIR, pdf_filename)
print(f"\n  Building PDF → {pdf_path}")

styles = getSampleStyleSheet()

def ps(name, parent='Normal', **kw):
    return ParagraphStyle(name, parent=styles[parent], **kw)

s_title    = ps('Title2',   fontName='Helvetica-Bold', fontSize=28, leading=34,
                textColor=HexColor(COLORS['primary']), spaceAfter=12, alignment=TA_CENTER)
s_h1       = ps('H1',       fontName='Helvetica-Bold', fontSize=18, leading=22,
                textColor=HexColor(COLORS['gray_dark']), spaceAfter=8, spaceBefore=10)
s_h2       = ps('H2',       fontName='Helvetica-Bold', fontSize=13, leading=16,
                textColor=HexColor(COLORS['gray_dark']), spaceAfter=6, spaceBefore=8)
s_body     = ps('Body',     fontName='Helvetica', fontSize=10,
                textColor=HexColor('#4b5563'), spaceAfter=6, leading=14)
s_caption  = ps('Caption',  fontName='Helvetica', fontSize=8,
                textColor=HexColor(COLORS['gray_medium']), spaceAfter=4, alignment=TA_CENTER)
s_subtitle = ps('Sub',      fontName='Helvetica', fontSize=14, leading=18,
                textColor=HexColor(COLORS['gray_medium']), spaceAfter=6, alignment=TA_CENTER)
s_center   = ps('Center',   fontName='Helvetica', fontSize=10,
                textColor=HexColor('#4b5563'), spaceAfter=6, alignment=TA_CENTER)

def table_style(n_rows, has_severity=False, severity_col=2):
    cmds = [
        ('BACKGROUND',  (0, 0), (-1, 0),  HexColor(COLORS['primary'])),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  HexColor(COLORS['white'])),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  9),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ALIGN',       (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0,0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',(0, 0), (-1, -1), 6),
        ('GRID',        (0, 0), (-1, -1), 0.5, HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),
         [HexColor(COLORS['white']), HexColor(COLORS['gray_light'])]),
    ]
    return TableStyle(cmds)

def apply_sev_colors(ts, data, sev_col=2):
    sev_map = {
        'High':   ('#fee2e2', COLORS['danger'],  'bold'),
        'Medium': ('#fef3c7', '#92400e',          'bold'),
        'Low':    ('#d1fae5', '#065f46',          'normal'),
    }
    for ri, row in enumerate(data[1:], start=1):
        cell = str(row[sev_col])
        if cell in sev_map:
            bg, fg, weight = sev_map[cell]
            ts.add('BACKGROUND', (sev_col, ri), (sev_col, ri), HexColor(bg))
            ts.add('TEXTCOLOR',  (sev_col, ri), (sev_col, ri), HexColor(fg))
            if weight == 'bold':
                ts.add('FONTNAME', (sev_col, ri), (sev_col, ri), 'Helvetica-Bold')
    return ts

# ── PRE-COMPUTE sorted issues + recommendations (used in summary & later pages) ─
sorted_issues = sorted(issues, key=lambda x: (-sev_score(x['severity']), -x['affected_rows']))

recs = []

def add_rec(priority, issue, recommendation, affected, impact):
    recs.append(dict(priority=priority, issue=issue,
                     recommendation=recommendation,
                     affected_columns=affected, impact=impact))

for iss in (i for i in sorted_issues if i['severity'] == 'High'):
    ct = iss['check_type']
    if ct == 'Missing Values':
        add_rec('P1 — Critical',
                f"Missing values in {iss['column']}",
                f"Impute (median/mode) or drop column if > 50% missing. "
                f"'{iss['column']}' has {iss['percentage']:.1f}% missing.",
                iss['column'],
                f"+{iss['percentage'] * 0.30:.1f} pts on Completeness")
    elif ct == 'Duplicate Rows':
        add_rec('P1 — Critical',
                "Exact duplicate rows present",
                f"Run df.drop_duplicates() to remove {iss['affected_rows']} duplicate rows "
                f"({iss['percentage']:.1f}%).",
                'ALL',
                f"+{iss['percentage'] * 0.15:.1f} pts on Uniqueness")
    elif ct == 'Type Validation':
        add_rec('P1 — Critical',
                f"Mixed types in {iss['column']}",
                f"Strip '$' and ',' from {iss['column']} then cast to float. {iss['finding']}",
                iss['column'],
                f"+{iss['percentage'] * 0.20:.1f} pts on Type Consistency")

seen_medium = set()
for iss in (i for i in sorted_issues if i['severity'] == 'Medium'):
    ct = iss['check_type']
    if ct in seen_medium:
        continue
    seen_medium.add(ct)
    if ct == 'Date Format Consistency':
        add_rec('P2 — High',
                f"Mixed date formats in {iss['column']}",
                f"Normalize to ISO 8601 (YYYY-MM-DD) using dateutil.parser or pd.to_datetime "
                f"with infer_datetime_format. {iss['finding']}",
                iss['column'], '+10 pts on Format Consistency')
    elif ct == 'Casing Inconsistency':
        add_rec('P2 — High',
                "Inconsistent text casing",
                "Apply .str.title() to name/category columns. "
                "Pick one convention and enforce it in the ingestion pipeline.",
                iss['column'], '+5 pts on Text Quality')
    elif ct == 'Whitespace Issues':
        add_rec('P2 — High',
                "Leading/trailing whitespace",
                "Apply .str.strip() across all string columns. "
                r"Use .str.replace(r'\s+', ' ', regex=True) for double spaces.",
                iss['column'], '+3 pts on Text Quality')
    elif ct == 'Column Name Quality':
        add_rec('P2 — High',
                "Unnamed / poorly-named columns",
                f"Rename or drop unnamed columns. Standardize all column names to snake_case. "
                f"{iss['finding']}",
                iss['column'], 'Improves maintainability')

if outlier_stats:
    worst = max(outlier_stats, key=lambda c: outlier_stats[c]['pct'])
    add_rec('P3 — Medium',
            f"Outliers in numeric columns (worst: {worst})",
            "Investigate outlier rows — validate against source data. "
            "Use IQR capping (Winsorization) for extreme values. "
            "Do NOT blindly drop — confirm business context first.",
            ', '.join(list(outlier_stats.keys())[:3]),
            f"+{avg_outlier_pct * 0.15:.1f} pts on Outlier Reasonability")

story = []

# ── PAGE 1: COVER ─────────────────────────────────────────────────────────────
story += [
    Spacer(1, 0.2 * inch),
    Paragraph("Data Quality Audit Report", s_title),
    Paragraph(dataset_name, s_subtitle),
    Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", s_center),
    Spacer(1, 0.15 * inch),
    RLImage(gauge_path, width=2.8 * inch, height=2.8 * inch, hAlign='CENTER'),
    Spacer(1, 0.12 * inch),
]
cover_data = [
    ['Metric', 'Value'],
    ['Total Rows',                   f'{total_rows:,}'],
    ['Total Columns',                str(total_cols)],
    ['Health Score',                 f'{health_score}/100  ({health_rating})'],
    ['Critical Issues (High)',       str(high_issues)],
    ['Warnings (Medium)',            str(medium_issues)],
    ['Info Items (Low)',             str(low_issues)],
    ['Columns with Missing Data',    str(sum(1 for v in missing_stats.values() if v['count'] > 0))],
    ['Duplicate Rows',               str(dup_count)],
]
t = Table(cover_data, colWidths=[3.2 * inch, 3.2 * inch])
t.setStyle(table_style(len(cover_data)))
story += [t, PageBreak()]

# ── PAGE 2: EXECUTIVE SUMMARY ─────────────────────────────────────────────────
story += [Paragraph("Executive Summary", s_h1)]
top3 = sorted(issues, key=lambda x: (-sev_score(x['severity']), -x['affected_rows']))[:3]
story += [
    Paragraph(
        f"This report presents a comprehensive data quality audit of <b>{dataset_name}</b>, "
        f"containing <b>{total_rows:,} rows</b> and <b>{total_cols} columns</b>. "
        f"The overall data health score is <b>{health_score}/100</b> — rated <b>{health_rating}</b>.",
        s_body,
    ),
    Paragraph(
        f"A total of <b>{high_issues} high-severity</b>, <b>{medium_issues} medium-severity</b>, "
        f"and <b>{low_issues} low-severity</b> issues were identified across "
        f"{len(set(i['check_type'] for i in issues))} quality-check categories.",
        s_body,
    ),
    Paragraph(
        "Key findings: " +
        "; ".join(f"<b>{i['column']}</b> — {i['finding']}" for i in top3) + ".",
        s_body,
    ),
    Paragraph(
        "Immediate attention should be given to high-severity issues (missing values, duplicates, "
        "type inconsistencies) which have the greatest impact on analytical reliability. "
        "Detailed findings and prioritized recommendations follow.",
        s_body,
    ),
    Spacer(1, 0.15 * inch),
]
score_data = [['Category', 'Weight', 'Score', 'Contribution']]
for cat in cat_labels:
    w = cat_weights[cat]
    s = category_scores[cat]
    score_data.append([cat, f'{w}%', f'{s:.1f}', f'{s * w / 100:.1f}'])
score_data.append(['TOTAL', '100%', '', f'{health_score}'])
t2 = Table(score_data, colWidths=[2.4 * inch, 0.9 * inch, 1 * inch, 1.2 * inch])
t2.setStyle(table_style(len(score_data)))
story += [t2, Spacer(1, 0.15 * inch),
          RLImage(pie_path, width=5.5 * inch, height=4.5 * inch, hAlign='CENTER'),
          Paragraph("Figure 3: Health Score Breakdown by Quality Category", s_caption),
          PageBreak()]

# ── PAGE 3: FINDINGS SUMMARY ──────────────────────────────────────────────────
story += [Paragraph("Findings Summary", s_h1)]

# Build sentences dynamically from audit results
_missing_cols  = [c for c, v in missing_stats.items() if v['count'] > 0]
_high_miss     = [c for c, v in missing_stats.items() if v['pct'] > 20]
_date_issues   = [i for i in issues if i['check_type'] == 'Date Format Consistency']
_type_issues   = [i for i in issues if i['check_type'] == 'Type Validation']
_casing_issues = [i for i in issues if i['check_type'] == 'Casing Inconsistency']
_ws_issues     = [i for i in issues if i['check_type'] == 'Whitespace Issues']
_col_issues    = [i for i in issues if i['check_type'] == 'Column Name Quality']
_p1_recs       = [r for r in recs if r['priority'].startswith('P1')]

_summary_sentences = []

# 1. Overall health
_summary_sentences.append(
    f"The dataset <b>{dataset_name}</b> contains <b>{total_rows:,} rows</b> and "
    f"<b>{total_cols} columns</b>, with an overall data health score of "
    f"<b>{health_score}/100</b>, rated <b>{health_rating}</b>."
)

# 2. Issue counts
_summary_sentences.append(
    f"The audit identified <b>{len(issues)} issues</b> in total — "
    f"<b>{high_issues} high-severity</b>, <b>{medium_issues} medium-severity</b>, "
    f"and <b>{low_issues} low-severity</b> — spanning "
    f"{len(set(i['check_type'] for i in issues))} quality-check categories."
)

# 3. Completeness
if _missing_cols:
    _summary_sentences.append(
        f"<b>Completeness</b> scored <b>{completeness:.1f}/100</b>: "
        f"{len(_missing_cols)} of {total_cols} columns contain missing values, "
        f"with an average missing rate of <b>{avg_missing_pct:.1f}%</b> across the dataset."
    )
    if _high_miss:
        _summary_sentences.append(
            f"Columns <b>{', '.join(_high_miss)}</b> exceed the 20% missing threshold "
            f"and require immediate imputation or removal."
        )
else:
    _summary_sentences.append(
        f"<b>Completeness</b> scored <b>{completeness:.1f}/100</b>: "
        f"no missing values were detected across any column."
    )

# 4. Uniqueness
if dup_count > 0:
    _summary_sentences.append(
        f"<b>Uniqueness</b> scored <b>{uniqueness:.1f}/100</b>: "
        f"<b>{dup_count} exact duplicate rows</b> ({dup_pct:.1f}%) were found and "
        f"should be removed before any downstream analysis."
    )
else:
    _summary_sentences.append(
        f"<b>Uniqueness</b> scored <b>{uniqueness:.1f}/100</b>: "
        f"no duplicate rows were detected."
    )

# 5. Type consistency
if _type_issues:
    _affected = ', '.join(i['column'] for i in _type_issues[:3])
    _summary_sentences.append(
        f"<b>Type Consistency</b> scored <b>{type_consistency:.1f}/100</b>: "
        f"mixed data types were found in <b>{_affected}</b>, "
        f"primarily due to currency-formatted strings that prevent numeric analysis."
    )
else:
    _summary_sentences.append(
        f"<b>Type Consistency</b> scored a strong <b>{type_consistency:.1f}/100</b> "
        f"with no type validation issues detected."
    )

# 6. Date format consistency
if _date_issues:
    _dc = _date_issues[0]
    _summary_sentences.append(
        f"<b>Format Consistency</b> scored <b>{format_consist:.1f}/100</b>: "
        f"column <b>{_dc['column']}</b> contains {_dc['finding']}, "
        f"which will cause parsing failures if not normalised to ISO 8601."
    )

# 7. Text quality
if _casing_issues or _ws_issues:
    _tq_cols = list({i['column'] for i in _casing_issues + _ws_issues})[:4]
    _summary_sentences.append(
        f"<b>Text Quality</b> scored <b>{text_quality:.1f}/100</b>: "
        f"inconsistent casing and/or leading/trailing whitespace were detected "
        f"in <b>{', '.join(_tq_cols)}</b>."
    )

# 8. Column naming
if _col_issues:
    _summary_sentences.append(
        f"<b>{len(_col_issues)} column(s)</b> have poor or auto-generated names "
        f"(e.g. <b>{_col_issues[0]['column']}</b>), which reduce maintainability "
        f"and should be renamed or dropped."
    )

# 9. Outliers
if outlier_stats:
    _worst_col = max(outlier_stats, key=lambda c: outlier_stats[c]['pct'])
    _summary_sentences.append(
        f"<b>Outlier Reasonability</b> scored <b>{outlier_reason:.1f}/100</b>: "
        f"statistical outliers were found in {len(outlier_stats)} numeric column(s), "
        f"with <b>{_worst_col}</b> being the most affected "
        f"({outlier_stats[_worst_col]['pct']:.1f}% of values outside IQR bounds)."
    )

# 10. Worst category
_worst_cat = min(category_scores, key=category_scores.get)
_summary_sentences.append(
    f"The lowest-scoring quality dimension is <b>{_worst_cat}</b> "
    f"at <b>{category_scores[_worst_cat]:.1f}/100</b>, "
    f"making it the primary focus area for remediation."
)

# 11. Top recommendation
if _p1_recs:
    _summary_sentences.append(
        f"The top priority action is: <b>{_p1_recs[0]['recommendation']}</b>"
    )

# 12. Closing
_summary_sentences.append(
    f"Addressing all {high_issues} high-severity issues is estimated to raise the health "
    f"score significantly; detailed remediation steps are provided in the Recommendations section."
)

for sent in _summary_sentences:
    story += [Paragraph(sent, s_body)]

story += [PageBreak()]

# ── PAGE 4: MISSING VALUES ────────────────────────────────────────────────────
story += [Paragraph("Missing Values Analysis", s_h1)]
n_miss_cols = sum(1 for v in missing_stats.values() if v['count'] > 0)
if missing_path:
    story += [
        Paragraph(
            f"Of {total_cols} columns, <b>{n_miss_cols}</b> contain missing values. "
            f"Average missing rate: <b>{avg_missing_pct:.1f}%</b>. "
            "Columns above 20% are flagged High and require imputation or removal.",
            s_body,
        ),
        Spacer(1, 0.1 * inch),
        RLImage(missing_path, width=6.5 * inch,
                height=min(8.5, max(3.0, n_miss_cols * 0.45)) * inch, hAlign='CENTER'),
        Paragraph("Figure 2: Missing values per column — red > 20%, orange 5–20%, green < 5%",
                  s_caption),
        Spacer(1, 0.15 * inch),
    ]
    miss_tbl = [['Column', 'Missing Count', 'Missing %', 'Severity']]
    for col, st in sorted(cols_with_missing.items(), key=lambda x: -x[1]['pct'])[:12]:
        sev = 'High' if st['pct'] > 20 else ('Medium' if st['pct'] >= 5 else 'Low')
        miss_tbl.append([col, str(st['count']), f"{st['pct']:.1f}%", sev])
    t3 = Table(miss_tbl, colWidths=[2.5 * inch, 1.2 * inch, 1.2 * inch, 1.1 * inch])
    ts3 = table_style(len(miss_tbl))
    ts3 = apply_sev_colors(ts3, miss_tbl, sev_col=3)
    t3.setStyle(ts3)
    story += [KeepTogether(t3)]
else:
    story += [Paragraph("No missing values detected.", s_body)]
story += [PageBreak()]

# ── PAGE 4: SEVERITY HEATMAP ──────────────────────────────────────────────────
story += [
    Paragraph("Issue Severity Overview", s_h1),
    Paragraph(
        "The heatmap maps each quality check (rows) against each column (columns). "
        "<b>H</b> = High severity, <b>M</b> = Medium, <b>L</b> = Low, blank = no issue.",
        s_body,
    ),
    Spacer(1, 0.1 * inch),
    RLImage(heatmap_path, width=7 * inch, height=3.8 * inch, hAlign='CENTER'),
    Paragraph("Figure 4: Issue Severity Heatmap", s_caption),
    Spacer(1, 0.15 * inch),
]
# Hotspots
hotspot_cols = {}
for issue in issues:
    for c in columns_to_show:
        if c in issue['column'] or issue['column'] == 'ALL':
            hotspot_cols[c] = hotspot_cols.get(c, 0) + sev_score(issue['severity'])
top_hot = sorted(hotspot_cols.items(), key=lambda x: -x[1])[:5]
story += [
    Paragraph(
        "Top issue hotspots: " +
        ", ".join(f"<b>{c}</b> (score {sc})" for c, sc in top_hot) + ".",
        s_body,
    ),
    PageBreak(),
]

# ── PAGE 5: DISTRIBUTIONS ─────────────────────────────────────────────────────
if hist_path:
    story += [
        Paragraph("Distribution Analysis", s_h1),
        Paragraph(
            f"Histograms for {len(hist_cols)} numeric column(s). "
            "Red dashed = median; orange dotted = IQR outlier boundaries (Q1 − 1.5×IQR, Q3 + 1.5×IQR). "
            "Values beyond orange lines are statistical outliers.",
            s_body,
        ),
        Spacer(1, 0.1 * inch),
        RLImage(hist_path, width=7 * inch, height=4.2 * inch, hAlign='CENTER'),
        Paragraph("Figure 5: Value distributions for numeric columns", s_caption),
        Spacer(1, 0.15 * inch),
    ]
    for col, st in outlier_stats.items():
        if st['count'] > 0:
            story += [Paragraph(
                f"<b>{col}</b>: {st['count']} outlier(s) ({st['pct']:.1f}%), "
                f"valid range [{st['lower']:.2f}, {st['upper']:.2f}].",
                s_body,
            )]
    story += [PageBreak()]

# ── PAGE 6: TREND ANALYSIS ────────────────────────────────────────────────────
if trend_path:
    story += [
        Paragraph("Trend Analysis", s_h1),
        Paragraph(
            f"Numeric values over time using <b>{primary_date_col}</b>. "
            "Raw data shown at low opacity; solid line = rolling average. "
            "Note: dates were parsed from mixed formats (MM/DD/YYYY, YYYY-MM-DD, DD-MM-YYYY).",
            s_body,
        ),
        Spacer(1, 0.1 * inch),
        RLImage(trend_path, width=7 * inch,
                height=min(7.0, 3.5 * len(trend_cols)) * inch, hAlign='CENTER'),
        Paragraph("Figure 6: Trends Over Time with rolling averages", s_caption),
        PageBreak(),
    ]
elif not date_columns:
    story += [Paragraph("No time-series data detected — trend page skipped.", s_body)]

# ── PAGE 7: DETAILED FINDINGS ─────────────────────────────────────────────────
story += [
    Paragraph("Detailed Findings", s_h1),
    Paragraph(
        f"All {len(issues)} issues sorted by severity (High → Low), then affected rows.",
        s_body,
    ),
    Spacer(1, 0.1 * inch),
]
ftbl = [['Column', 'Check Type', 'Severity', 'Finding', 'Rows', '%']]
for iss in sorted_issues:
    ftbl.append([
        iss['column'][:22],
        iss['check_type'][:22],
        iss['severity'],
        iss['finding'][:55],
        str(iss['affected_rows']),
        f"{iss['percentage']:.1f}%",
    ])
t4 = Table(ftbl,
           colWidths=[1.2*inch, 1.35*inch, 0.72*inch, 2.6*inch, 0.58*inch, 0.5*inch],
           repeatRows=1)
ts4 = table_style(len(ftbl))
ts4 = apply_sev_colors(ts4, ftbl, sev_col=2)
t4.setStyle(ts4)
story += [t4]

# ── PAGE 8: RECOMMENDATIONS ───────────────────────────────────────────────────
# Recommendations table — landscape page, Paragraph cells for word wrap
LETTER_P = letter
LETTER_L = ls_pagesize(letter)   # (792, 612)
s_cell = ps('Cell', fontName='Helvetica', fontSize=8,
            textColor=HexColor('#374151'), leading=11, spaceAfter=0)
s_cell_bold = ps('CellBold', fontName='Helvetica-Bold', fontSize=8,
                 textColor=HexColor('#374151'), leading=11, spaceAfter=0)

# Switch to landscape before the recommendations page
story += [NextPageTemplate('Landscape'), PageBreak()]
story += [Paragraph("Recommendations", s_h1)]

rec_tbl = [['Priority', 'Issue', 'Recommendation', 'Affected Columns', 'Est. Impact']]
for r in recs:
    rec_tbl.append([
        Paragraph(r['priority'],           s_cell_bold),
        Paragraph(r['issue'],              s_cell),
        Paragraph(r['recommendation'],     s_cell),
        Paragraph(r['affected_columns'],   s_cell),
        Paragraph(r['impact'],             s_cell),
    ])
# Landscape available width: 11" - 1.5" margins = 9.5"
t5 = Table(rec_tbl,
           colWidths=[1.1*inch, 1.6*inch, 4.4*inch, 1.3*inch, 1.1*inch],
           repeatRows=1, splitByRow=1)
ts5 = table_style(len(rec_tbl))
t5.setStyle(ts5)
story += [t5]

# ── BUILD PDF with portrait + landscape templates ─────────────────────────────
def on_page_portrait(canvas, doc):
    canvas.saveState()
    if doc.page > 1:
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(HexColor(COLORS['gray_medium']))
        canvas.drawString(0.75 * inch, 0.4 * inch, "Data Quality Audit Report")
        canvas.drawRightString(LETTER_P[0] - 0.75 * inch, 0.4 * inch, f"Page {doc.page}")
    canvas.restoreState()

def on_page_landscape(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(HexColor(COLORS['gray_medium']))
    canvas.drawString(0.75 * inch, 0.4 * inch, "Data Quality Audit Report")
    canvas.drawRightString(LETTER_L[0] - 0.75 * inch, 0.4 * inch, f"Page {doc.page}")
    canvas.restoreState()

portrait_frame  = Frame(0.75*inch, 0.75*inch,
                        LETTER_P[0] - 1.5*inch, LETTER_P[1] - 1.5*inch,
                        id='portrait_body')
landscape_frame = Frame(0.75*inch, 0.75*inch,
                        LETTER_L[0] - 1.5*inch, LETTER_L[1] - 1.5*inch,
                        id='landscape_body')

portrait_tpl  = PageTemplate(id='Portrait',  frames=[portrait_frame],
                              onPage=on_page_portrait,  pagesize=LETTER_P)
landscape_tpl = PageTemplate(id='Landscape', frames=[landscape_frame],
                              onPage=on_page_landscape, pagesize=LETTER_L)

doc = BaseDocTemplate(pdf_path,
                      pageTemplates=[portrait_tpl, landscape_tpl],
                      pagesize=LETTER_P)
doc.build(story)
print(f"  PDF saved ({os.path.getsize(pdf_path) // 1024} KB)")

# ── STEP 6: EXCEL WORKBOOK ────────────────────────────────────────────────────
xlsx_filename = f"audit_data_{dataset_name}_{today}.xlsx"
xlsx_path     = os.path.join(OUTPUT_DIR, xlsx_filename)
print(f"\n  Building Excel → {xlsx_path}")

wb = openpyxl.Workbook()

_BLUE_FILL   = PatternFill('solid', fgColor='FF1a56db')
_WHITE_FONT  = Font(bold=True, color='FFFFFFFF', size=10)
_ALT_FILL    = PatternFill('solid', fgColor='FFF3F4F6')
_THIN_BORDER = Border(
    left=Side(style='thin', color='FFD1D5DB'),
    right=Side(style='thin', color='FFD1D5DB'),
    top=Side(style='thin', color='FFD1D5DB'),
    bottom=Side(style='thin', color='FFD1D5DB'),
)
_SEV_FILLS   = {
    'High':   PatternFill('solid', fgColor='FFFEE2E2'),
    'Medium': PatternFill('solid', fgColor='FFFEF3C7'),
    'Low':    PatternFill('solid', fgColor='FFD1FAE5'),
}
_SEV_FONTS = {
    'High':   Font(bold=True, color='FFEF4444', size=9),
    'Medium': Font(bold=True, color='FF92400E', size=9),
    'Low':    Font(bold=True, color='FF065F46', size=9),
}

def hdr(cell, label):
    cell.value      = label
    cell.font       = _WHITE_FONT
    cell.fill       = _BLUE_FILL
    cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border     = _THIN_BORDER

def dcell(cell, value, row_idx):
    cell.value     = value
    cell.fill      = _ALT_FILL if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFFFF')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border    = _THIN_BORDER

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Summary'
for col_letter, label in (('A', 'Metric'), ('B', 'Value')):
    hdr(ws1[f'{col_letter}1'], label)
ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 28

summary_rows = [
    ('Dataset Name',                    dataset_name),
    ('Total Rows',                       total_rows),
    ('Total Columns',                    total_cols),
    ('Health Score',                     f'{health_score}/100'),
    ('Health Rating',                    health_rating),
    ('Critical Issues (High)',           high_issues),
    ('Warnings (Medium)',                medium_issues),
    ('Info Items (Low)',                 low_issues),
    ('Columns with Missing Data',        sum(1 for v in missing_stats.values() if v['count'] > 0)),
    ('Duplicate Rows',                   dup_count),
    ('Date Generated',                   datetime.now().strftime('%Y-%m-%d %H:%M')),
]
for i, (metric, value) in enumerate(summary_rows, start=2):
    dcell(ws1[f'A{i}'], metric, i)
    dcell(ws1[f'B{i}'], value, i)
    ws1[f'A{i}'].font = Font(bold=True, size=10)
ws1.freeze_panes = 'A2'

# ── Sheet 2: Details ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Details')
det_hdrs  = ['Column Name', 'Check Type', 'Severity', 'Finding',
             'Affected Rows', 'Percentage', 'Example Values']
det_widths = [20, 24, 11, 52, 14, 12, 32]
for j, (h, w) in enumerate(zip(det_hdrs, det_widths), 1):
    hdr(ws2.cell(row=1, column=j), h)
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

for i, iss in enumerate(sorted_issues, 2):
    row_vals = [iss['column'], iss['check_type'], iss['severity'], iss['finding'],
                iss['affected_rows'], f"{iss['percentage']:.1f}%", iss['example_values']]
    for j, v in enumerate(row_vals, 1):
        dcell(ws2.cell(row=i, column=j), v, i)
    sev = iss['severity']
    sev_cell = ws2.cell(row=i, column=3)
    sev_cell.fill = _SEV_FILLS.get(sev, _ALT_FILL)
    sev_cell.font = _SEV_FONTS.get(sev, Font(size=9))

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = ws2.dimensions

# ── Sheet 3: Recommendations ──────────────────────────────────────────────────
ws3 = wb.create_sheet('Recommendations')
rec_hdrs   = ['Priority', 'Issue', 'Recommendation', 'Affected Columns', 'Estimated Impact']
rec_widths = [16, 32, 62, 22, 26]
for j, (h, w) in enumerate(zip(rec_hdrs, rec_widths), 1):
    hdr(ws3.cell(row=1, column=j), h)
    ws3.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

for i, r in enumerate(recs, 2):
    row_vals = [r['priority'], r['issue'], r['recommendation'],
                r['affected_columns'], r['impact']]
    for j, v in enumerate(row_vals, 1):
        c = ws3.cell(row=i, column=j, value=v)
        c.fill      = _ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFFFF')
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border    = _THIN_BORDER
    ws3.row_dimensions[i].height = 50

ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = ws3.dimensions

# ── Sheet 4: Severity Heatmap ──────────────────────────────────────────────────
ws4 = wb.create_sheet('Severity Heatmap')

_HM_FILLS = {
    0: PatternFill('solid', fgColor='FFF3F4F6'),  # none
    1: PatternFill('solid', fgColor='FFD1FAE5'),  # Low  — green
    2: PatternFill('solid', fgColor='FFFEF3C7'),  # Medium — amber
    3: PatternFill('solid', fgColor='FFFEE2E2'),  # High  — red
}
_HM_FONTS = {
    0: Font(size=9, color='FF9CA3AF'),
    1: Font(bold=True, size=9, color='FF065F46'),
    2: Font(bold=True, size=9, color='FF92400E'),
    3: Font(bold=True, size=9, color='FFEF4444'),
}

# Header row: column names starting at B1
ws4.column_dimensions['A'].width = 26
ws4.row_dimensions[1].height = 30
for cj, col in enumerate(columns_to_show, start=2):
    col_letter = openpyxl.utils.get_column_letter(cj)
    c = ws4.cell(row=1, column=cj, value=col)
    c.fill      = _BLUE_FILL
    c.font      = _WHITE_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = _THIN_BORDER
    ws4.column_dimensions[col_letter].width = 14

# Top-left corner cell (A1)
corner = ws4.cell(row=1, column=1, value='Check Type \\ Column')
corner.fill      = _BLUE_FILL
corner.font      = Font(bold=True, color='FFFFFFFF', size=9)
corner.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
corner.border    = _THIN_BORDER

# Row headers and heatmap cells
for ri, check in enumerate(check_types, start=2):
    lbl = ws4.cell(row=ri, column=1, value=check)
    lbl.fill      = _BLUE_FILL
    lbl.font      = _WHITE_FONT
    lbl.alignment = Alignment(vertical='center', wrap_text=True)
    lbl.border    = _THIN_BORDER
    ws4.row_dimensions[ri].height = 22

    for cj in range(len(columns_to_show)):
        score = int(sev_matrix[ri - 2, cj])
        label = sev_labels[ri - 2, cj]
        cell  = ws4.cell(row=ri, column=cj + 2, value=label if label else '')
        cell.fill      = _HM_FILLS[score]
        cell.font      = _HM_FONTS[score]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _THIN_BORDER

# Freeze the header row and label column
ws4.freeze_panes = 'B2'

wb.save(xlsx_path)
print(f"  Excel saved ({os.path.getsize(xlsx_path) // 1024} KB)")

# ── CLEANUP ────────────────────────────────────────────────────────────────────
shutil.rmtree(charts_dir, ignore_errors=True)

# ── FINAL SUMMARY ─────────────────────────────────────────────────────────────
print(f"\n{'═'*60}")
print(f"  AUDIT COMPLETE")
print(f"{'═'*60}")
print(f"  Health Score : {health_score}/100  ({health_rating})")
print(f"  Issues       : {high_issues} High | {medium_issues} Medium | {low_issues} Low")
print(f"\n  Top 3 issues:")
for k, iss in enumerate(top3, 1):
    print(f"    {k}. [{iss['severity']}] {iss['column']}: {iss['finding']}")
print(f"\n  #1 Recommendation:")
if recs:
    print(f"    {recs[0]['priority']}: {recs[0]['recommendation']}")
print(f"\n  PDF  → {pdf_path}")
print(f"  XLSX → {xlsx_path}")
print(f"{'═'*60}\n")
